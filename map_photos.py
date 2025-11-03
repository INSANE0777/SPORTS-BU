import openpyxl
import os
from pathlib import Path

def extract_players_from_excel(excel_path):
    """
    Extract all players with their names and other details from Excel.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Find the column indices
    header_row = 1
    name_col = None
    unique_id_col = None
    photo_col = None
    
    for cell in ws[header_row]:
        if cell.value:
            cell_value_lower = str(cell.value).strip().lower()
            if cell_value_lower == "name":
                name_col = cell.column
            elif cell_value_lower == "uniqueid" or cell_value_lower == "enrollment":
                unique_id_col = cell.column
            elif cell_value_lower == "photo":
                photo_col = cell.column
    
    if name_col is None:
        print("Error: 'name' column not found")
        return []
    
    players = []
    
    # Iterate through all rows (skip header)
    for row_num in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row_num, column=name_col)
        if name_cell and name_cell.value:
            name = str(name_cell.value).strip()
            
            unique_id = None
            if unique_id_col:
                unique_id_cell = ws.cell(row=row_num, column=unique_id_col)
                if unique_id_cell and unique_id_cell.value:
                    unique_id = str(unique_id_cell.value).strip()
            
            # Try to extract photo link for reference
            photo_link = None
            if photo_col:
                photo_cell = ws.cell(row=row_num, column=photo_col)
                if photo_cell:
                    if photo_cell.hyperlink:
                        photo_link = photo_cell.hyperlink.target or photo_cell.hyperlink.location
                    elif photo_cell.value and isinstance(photo_cell.value, str):
                        photo_link = photo_cell.value
            
            players.append({
                'name': name,
                'unique_id': unique_id,
                'photo_link': photo_link,
                'row': row_num
            })
    
    wb.close()
    return players

def list_public_photos():
    """
    List all photo files in the public folder.
    """
    public_dir = Path("public")
    photos = []
    
    # Common image extensions
    image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.heic', '.heif', '.pdf'}
    
    for file in public_dir.iterdir():
        if file.is_file() and file.suffix.lower() in image_extensions:
            # Skip favicons and logos
            if any(skip in file.name.lower() for skip in ['favicon', 'logo', 'bennett', 'sports', 'placeholder', 'site']):
                continue
            photos.append(file.name)
    
    return photos

def normalize_name(name):
    """
    Normalize name for matching (remove extra spaces, case insensitive).
    """
    if not name:
        return ""
    # Remove extra spaces and convert to lowercase
    return " ".join(name.lower().split())

def find_matching_photo(player_name, photo_files):
    """
    Find matching photo for a player name.
    """
    normalized_player = normalize_name(player_name)
    
    for photo in photo_files:
        # Remove extension and normalize
        photo_base = Path(photo).stem.lower()
        normalized_photo = normalize_name(photo_base)
        
        # Check if player name matches photo name
        if normalized_player == normalized_photo:
            return photo
        
        # Also check if photo name is a substring of player name or vice versa
        # (useful for handling middle names, titles, etc.)
        if normalized_player in normalized_photo or normalized_photo in normalized_player:
            return photo
    
    return None

def main():
    excel_file = "Auction Players.xlsx"
    
    print("Reading Excel file...")
    players = extract_players_from_excel(excel_file)
    print(f"Found {len(players)} players in Excel")
    
    print("\nListing photos in public folder...")
    photos = list_public_photos()
    print(f"Found {len(photos)} photos in public folder")
    
    print("\n" + "="*80)
    print("MATCHING PLAYERS WITH PHOTOS")
    print("="*80)
    
    matched = []
    unmatched_players = []
    
    for player in players:
        photo = find_matching_photo(player['name'], photos)
        if photo:
            photo_path = f"/{photo}"  # Public folder files are served from root
            matched.append({
                'player': player,
                'photo': photo_path
            })
            print(f"✓ {player['name']} -> {photo}")
        else:
            unmatched_players.append(player)
            print(f"✗ {player['name']} -> NO MATCH")
    
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Total players: {len(players)}")
    print(f"Matched: {len(matched)}")
    print(f"Unmatched: {len(unmatched_players)}")
    
    print("\n" + "="*80)
    print("UNMATCHED PLAYERS")
    print("="*80)
    for player in unmatched_players:
        print(f"  - {player['name']} (Row {player['row']})")
    
    # Generate update script
    print("\n" + "="*80)
    print("GENERATING UPDATE COMMANDS")
    print("="*80)
    
    print("\n# Generated photo update commands")
    print("# These can be used to update Appwrite database")
    
    for match in matched:
        name = match['player']['name']
        unique_id = match['player'].get('unique_id', 'UNKNOWN')
        photo = match['photo']
        
        print(f"\n# Player: {name} ({unique_id})")
        print(f"# Photo: {photo}")
        print(f"# Update photo field in Appwrite for player with uniqueId: {unique_id}")
        print(f"# OR update by name: {name}")
    
    # Save to file
    output_file = "photo_mappings.txt"
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("PLAYER PHOTO MAPPINGS\n")
        f.write("="*80 + "\n\n")
        
        f.write("MATCHED PLAYERS:\n")
        f.write("-"*80 + "\n")
        for match in matched:
            name = match['player']['name']
            unique_id = match['player'].get('unique_id', 'N/A')
            photo = match['photo']
            f.write(f"{name} ({unique_id}) -> {photo}\n")
        
        f.write("\n\nUNMATCHED PLAYERS:\n")
        f.write("-"*80 + "\n")
        for player in unmatched_players:
            f.write(f"{player['name']} (Row {player['row']})\n")
    
    print(f"\n✓ Mappings saved to {output_file}")

if __name__ == "__main__":
    main()

